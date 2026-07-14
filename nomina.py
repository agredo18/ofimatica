"""
Automatizacion de nomina mensual - Empresa colombiana
=====================================================
Genera un archivo Excel (.xlsx) con la liquidacion de nomina de la plantilla,
un resumen gerencial basado en formulas de Excel, e imprime un analisis de
resultados por consola.

Requiere: Python 3.7+ y OpenPyXL.
Uso:      python nomina.py
"""

from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# SECCION 1: PARAMETROS Y REGLAS DE NEGOCIO
# ---------------------------------------------------------------------------

# Tope salarial que da derecho al auxilio de transporte (2 SMMLV aprox.).
TOPE_AUXILIO_TRANSPORTE = 2_847_000
VALOR_AUXILIO_TRANSPORTE = 200_000

HORAS_MES = 240            # Jornada mensual usada para hallar el valor hora
RECARGO_HORA_EXTRA = 1.25  # La hora extra vale 25% mas que la hora ordinaria

PORCENTAJE_SALUD = 0.04    # Aporte del empleado a salud
PORCENTAJE_PENSION = 0.04  # Aporte del empleado a pension

ARCHIVO_SALIDA = "nomina_empresa.xlsx"

MESES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]

# ---------------------------------------------------------------------------
# SECCION 2: DATOS DE ENTRADA (10 empleados)
# ---------------------------------------------------------------------------

EMPLEADOS = [
    # (Documento, Nombre, Cargo, Salario base, Horas extras)
    (1032456789, "Laura Gomez Restrepo",    "Desarrollador Junior",  2_300_000, 12),
    (1015879234, "Andres Muñoz Cardenas",   "Desarrollador Senior",  4_500_000,  8),
    (1098234567, "Camila Rojas Vargas",     "Analista de Datos",     3_200_000, 10),
    (1004567812, "Julian Ospina Herrera",   "Auxiliar Contable",     2_100_000, 20),
    (1023458796, "Valentina Torres Diaz",   "Diseñadora UX",         2_800_000,  6),
    (1078965412, "Sebastian Rios Mejia",    "Gerente de Proyecto",   5_000_000,  0),
    (1045678923, "Daniela Castro Pineda",   "Analista de Calidad",   2_650_000, 15),
    (1067894523, "Mateo Salazar Quintero",  "Ingeniero DevOps",      4_100_000,  9),
    (1089452367, "Isabella Nieto Ramirez",  "Asistente Administrativa", 2_000_000, 18),
    (1056783412, "Santiago Bedoya Lopez",   "Lider Tecnico",         4_800_000,  5),
]

# ---------------------------------------------------------------------------
# SECCION 3: CALCULOS DE NOMINA
# ---------------------------------------------------------------------------


def liquidar_empleado(documento, nombre, cargo, salario, horas_extras):
    """Aplica las reglas de negocio a un empleado y devuelve su liquidacion.

    Los valores monetarios se redondean a 2 decimales para que los totales de
    Excel y los del reporte por consola coincidan exactamente.
    """
    auxilio = VALOR_AUXILIO_TRANSPORTE if salario <= TOPE_AUXILIO_TRANSPORTE else 0

    valor_hora = round(salario / HORAS_MES, 2)
    valor_hora_extra = round(valor_hora * RECARGO_HORA_EXTRA, 2)
    total_horas_extras = round(horas_extras * valor_hora_extra, 2)

    devengado = round(salario + auxilio + total_horas_extras, 2)

    salud = round(devengado * PORCENTAJE_SALUD, 2)
    pension = round(devengado * PORCENTAJE_PENSION, 2)
    neto = round(devengado - salud - pension, 2)

    return {
        "documento": documento,
        "nombre": nombre,
        "cargo": cargo,
        "salario": salario,
        "horas_extras": horas_extras,
        "auxilio": auxilio,
        "valor_hora": valor_hora,
        "valor_hora_extra": valor_hora_extra,
        "total_horas_extras": total_horas_extras,
        "devengado": devengado,
        "salud": salud,
        "pension": pension,
        "neto": neto,
    }


def liquidar_nomina(empleados):
    """Liquida la lista completa de empleados."""
    return [liquidar_empleado(*empleado) for empleado in empleados]


# ---------------------------------------------------------------------------
# SECCION 4: ESTILOS REUTILIZABLES
# ---------------------------------------------------------------------------

AZUL_CORPORATIVO = "1F4E78"
FORMATO_MONEDA = "#,##0"
ANCHO_COLUMNA = 16

fuente_encabezado = Font(bold=True, color="FFFFFF", size=11)
relleno_encabezado = PatternFill("solid", start_color=AZUL_CORPORATIVO, end_color=AZUL_CORPORATIVO)
alineacion_encabezado = Alignment(horizontal="center", vertical="center", wrap_text=True)

_linea = Side(style="thin", color="B4C6E7")
borde_celda = Border(left=_linea, right=_linea, top=_linea, bottom=_linea)


def aplicar_estilo_encabezado(hoja, fila, cantidad_columnas):
    """Pinta de azul, en negrita y en blanco la fila de encabezados."""
    for columna in range(1, cantidad_columnas + 1):
        celda = hoja.cell(row=fila, column=columna)
        celda.font = fuente_encabezado
        celda.fill = relleno_encabezado
        celda.alignment = alineacion_encabezado
        celda.border = borde_celda


def ajustar_ancho_columnas(hoja, cantidad_columnas, ancho=ANCHO_COLUMNA):
    """Fija un ancho uniforme para las columnas de la hoja."""
    for columna in range(1, cantidad_columnas + 1):
        hoja.column_dimensions[get_column_letter(columna)].width = ancho


# ---------------------------------------------------------------------------
# SECCION 5: HOJA 1 - DETALLE DE NOMINA
# ---------------------------------------------------------------------------

ENCABEZADOS_NOMINA = [
    "Documento", "Nombre", "Cargo", "Salario", "Horas Extras",
    "Auxilio Transporte", "Valor Hora", "Valor Hora Extra",
    "Total Horas Extras", "Devengado", "Salud", "Pensión", "Neto a Pagar",
]

# Columnas (1-based) que llevan formato de moneda con separador de miles.
COLUMNAS_MONETARIAS = (4, 6, 7, 8, 9, 10, 11, 12, 13)


def construir_hoja_nomina(libro, nomina, titulo):
    """Crea y llena la hoja de detalle de nomina."""
    hoja = libro.active
    hoja.title = titulo

    # Fila de encabezados
    hoja.append(ENCABEZADOS_NOMINA)
    aplicar_estilo_encabezado(hoja, 1, len(ENCABEZADOS_NOMINA))

    # Filas de empleados
    for registro in nomina:
        hoja.append([
            registro["documento"],
            registro["nombre"],
            registro["cargo"],
            registro["salario"],
            registro["horas_extras"],
            registro["auxilio"],
            registro["valor_hora"],
            registro["valor_hora_extra"],
            registro["total_horas_extras"],
            registro["devengado"],
            registro["salud"],
            registro["pension"],
            registro["neto"],
        ])

    # Formato de las celdas de datos
    ultima_fila = 1 + len(nomina)
    for fila in range(2, ultima_fila + 1):
        for columna in range(1, len(ENCABEZADOS_NOMINA) + 1):
            celda = hoja.cell(row=fila, column=columna)
            celda.border = borde_celda
            if columna in COLUMNAS_MONETARIAS:
                celda.number_format = FORMATO_MONEDA
                celda.alignment = Alignment(horizontal="right")
            elif columna in (1, 5):  # Documento y Horas Extras
                celda.alignment = Alignment(horizontal="center")

    ajustar_ancho_columnas(hoja, len(ENCABEZADOS_NOMINA))
    hoja.freeze_panes = "A2"  # Encabezado siempre visible al desplazarse
    return hoja


# ---------------------------------------------------------------------------
# SECCION 6: HOJA 2 - RESUMEN GERENCIAL (con formulas de Excel)
# ---------------------------------------------------------------------------


def construir_hoja_resumen(libro, titulo_nomina, cantidad_empleados):
    """Crea la hoja de indicadores.

    Los valores NO se calculan en Python: se escriben formulas que referencian
    la hoja de nomina, de modo que el resumen se recalcula solo si el detalle
    cambia dentro de Excel.
    """
    hoja = libro.create_sheet("Resumen Gerencial")

    fila_inicial, fila_final = 2, 1 + cantidad_empleados
    # El nombre de la hoja lleva espacios y tilde -> debe ir entre comillas simples.
    ref = f"'{titulo_nomina}'"

    def rango(columna):
        return f"{ref}!{columna}{fila_inicial}:{columna}{fila_final}"

    indicadores = [
        ("Total Empleados",    f"=COUNTA({rango('A')})",   "0"),
        ("Total Nómina",       f"=SUM({rango('J')})",      FORMATO_MONEDA),
        ("Total Salud",        f"=SUM({rango('K')})",      FORMATO_MONEDA),
        ("Total Pensión",      f"=SUM({rango('L')})",      FORMATO_MONEDA),
        ("Mayor Salario",      f"=MAX({rango('D')})",      FORMATO_MONEDA),
        ("Menor Salario",      f"=MIN({rango('D')})",      FORMATO_MONEDA),
        ("Promedio Salarial",  f"=AVERAGE({rango('D')})",  FORMATO_MONEDA),
    ]

    hoja.append(["Indicador", "Valor"])
    aplicar_estilo_encabezado(hoja, 1, 2)

    for fila, (nombre, formula, formato) in enumerate(indicadores, start=2):
        celda_indicador = hoja.cell(row=fila, column=1, value=nombre)
        celda_valor = hoja.cell(row=fila, column=2, value=formula)

        celda_indicador.font = Font(bold=True)
        celda_indicador.border = borde_celda
        celda_valor.number_format = formato
        celda_valor.alignment = Alignment(horizontal="right")
        celda_valor.border = borde_celda

    ajustar_ancho_columnas(hoja, 2, ancho=22)
    return hoja


# ---------------------------------------------------------------------------
# SECCION 7: REPORTE POR CONSOLA
# ---------------------------------------------------------------------------


def formato_pesos(valor):
    """Formatea un numero como moneda colombiana: $ 1.234.567"""
    return "$ " + f"{valor:,.0f}".replace(",", ".")


def imprimir_analisis(nomina, titulo_nomina):
    """Imprime el analisis de resultados solicitado por la gerencia."""
    total_devengado = sum(e["devengado"] for e in nomina)
    total_salud = sum(e["salud"] for e in nomina)
    total_pension = sum(e["pension"] for e in nomina)
    mejor_pagado = max(nomina, key=lambda e: e["neto"])

    ancho = 68
    print()
    print("=" * ancho)
    print(f"  ANALISIS DE RESULTADOS - {titulo_nomina.upper()}".center(ancho))
    print("=" * ancho)

    print(f"\n1. Valor total pagado en nomina (devengado) : {formato_pesos(total_devengado)}")
    print(f"2. Dinero descontado por salud (4%)         : {formato_pesos(total_salud)}")
    print(f"3. Dinero descontado por pension (4%)       : {formato_pesos(total_pension)}")
    print(f"   Total descuentos                         : {formato_pesos(total_salud + total_pension)}")
    print(f"   Total neto a pagar                       : "
          f"{formato_pesos(total_devengado - total_salud - total_pension)}")

    print("\n4. Empleado con el mayor pago neto:")
    print(f"   Nombre     : {mejor_pagado['nombre']}")
    print(f"   Cargo      : {mejor_pagado['cargo']}")
    print(f"   Documento  : {mejor_pagado['documento']}")
    print(f"   Neto a pagar: {formato_pesos(mejor_pagado['neto'])}")

    print("\n5. Ventajas de automatizar la nomina con Python:")
    ventajas = [
        "Elimina los errores de digitacion y de formulas mal copiadas en Excel.",
        "Liquida los 10 empleados (o 10.000) en segundos, no en horas.",
        "Las reglas de negocio quedan en un solo lugar: cambiar el % de salud o",
        "  el auxilio de transporte es modificar una constante, no cada celda.",
        "El proceso es reproducible y auditable: mismos datos, mismo resultado.",
        "Genera reportes con formato profesional listos para la gerencia.",
        "Libera al area de talento humano de tareas repetitivas.",
    ]
    for ventaja in ventajas:
        print(f"   - {ventaja}" if not ventaja.startswith("  ") else f"    {ventaja.strip()}")

    print("\n" + "=" * ancho)


# ---------------------------------------------------------------------------
# SECCION 8: PROGRAMA PRINCIPAL
# ---------------------------------------------------------------------------


def main():
    hoy = date.today()
    titulo_nomina = f"Nómina {MESES[hoy.month - 1]} {hoy.year}"

    # 1) Aplicar reglas de negocio
    nomina = liquidar_nomina(EMPLEADOS)

    # 2) Construir el libro de Excel
    libro = Workbook()
    construir_hoja_nomina(libro, nomina, titulo_nomina)
    construir_hoja_resumen(libro, titulo_nomina, len(nomina))

    # 3) Guardar
    try:
        libro.save(ARCHIVO_SALIDA)
    except PermissionError:
        print(f"\n[ERROR] No se pudo guardar '{ARCHIVO_SALIDA}'.")
        print("        Cierra el archivo en Excel y vuelve a ejecutar el programa.")
        return

    print(f"\n[OK] Archivo generado: {ARCHIVO_SALIDA}")
    print(f"     Hoja 1: {titulo_nomina}")
    print("     Hoja 2: Resumen Gerencial")

    # 4) Analisis por consola
    imprimir_analisis(nomina, titulo_nomina)


if __name__ == "__main__":
    main()
